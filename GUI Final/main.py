import tkinter as tk  # Import tkinter library for GUI
from logging import root
from tkinter import messagebox  # Import messagebox for displaying messages

import typing_test  # Import the TypingTestWindow class from typing_test module
import stopwatch_test  # Import the StopwatchTestWindow class from stopwatch_test module
import os  # Import os module for file operations
import openpyxl  # Import openpyxl library for working with Excel files
from openpyxl import Workbook  # Import Workbook class from openpyxl
from openpyxl.utils import get_column_letter  # Utility function to get column letter from column index
# pip install pillow
from PIL import Image, ImageTk, ImageDraw

# Define color constants from the palette
BG_COLOR = "#D3BDB0"
BUTTON_COLOR = "#89937C"
TEXT_COLOR = "white"


# Define the main application class
class TypingTestApp(tk.Tk):
    def __init__(self):
        super().__init__()  # Initialize the tkinter application
        self.title("Typing Test App")  # Set window title
        self.geometry("600x500")  # Set window size
        self.configure(bg=BG_COLOR)  # Set background color

        # Load images
        keyboard_image = Image.open("keyboard.png")
        keyboard_photo = ImageTk.PhotoImage(keyboard_image)

        # Create labels for images
        keyboard_label = tk.Label(self, image=keyboard_photo, bg=BG_COLOR)
        keyboard_label.image = keyboard_photo  # Keep a reference to avoid garbage collection
        keyboard_label.pack(pady=0.5)

        # Load images
        computer_image = Image.open("output-onlinepngtools.png")
        resized = computer_image.resize((200, 200))
        new_pic = ImageTk.PhotoImage(resized)

        # Create labels for images
        computer_label = tk.Label(self, image=new_pic, bg=BG_COLOR)
        computer_label.image = new_pic  # Keep a reference to avoid garbage collection
        computer_label.pack(pady=0.01)


        # Create buttons for Typing Test
        button1 = tk.Button(self, text="One-Minute Typing Test", command=self.open_typing_test, bg=BUTTON_COLOR,
                            fg=TEXT_COLOR)
        button1.pack(pady=10)  # Pack the button into the window

        record_button1 = tk.Button(self, text="Record One-Minute Typing Test", command=self.record_one_minute_test,
                                   bg=BUTTON_COLOR, fg=TEXT_COLOR)
        record_button1.pack(pady=5)  # Pack the button into the window

        # Create buttons for Stopwatch Test
        button2 = tk.Button(self, text="Stopwatch Timing Test", command=self.open_stopwatch_test, bg=BUTTON_COLOR,
                            fg=TEXT_COLOR)
        button2.pack(pady=10)  # Pack the button into the window

        record_button2 = tk.Button(self, text="Record Stopwatch Timing Test", command=self.record_stopwatch_test,
                                   bg=BUTTON_COLOR, fg=TEXT_COLOR)
        record_button2.pack(pady=5)  # Pack the button into the window


    # Method to open the One-Minute Typing Test window
    def open_typing_test(self):
        typing_test_app = typing_test.TypingTestWindow(self)
        typing_test_app.mainloop()  # Start the typing test window

    # Method to open the Stopwatch Timing Test window
    def open_stopwatch_test(self):
        stopwatch_test_app = stopwatch_test.StopwatchTestWindow(self)
        stopwatch_test_app.mainloop()  # Start the stopwatch test window

    # Method to open the window for recording results of One-Minute Typing Test
    def record_one_minute_test(self):
        record_window = RecordTestWindow(self, test_type="OneMinute")
        record_window.mainloop()  # Start the record window

    # Method to open the window for recording results of Stopwatch Timing Test
    def record_stopwatch_test(self):
        record_window = RecordTestWindow(self, test_type="Stopwatch")
        record_window.mainloop()  # Start the record window


# Class for the window to record test results
class RecordTestWindow(tk.Toplevel):
    def __init__(self, parent, test_type):
        super().__init__(parent)  # Initialize the Toplevel window
        self.parent = parent  # Store reference to parent window
        self.title("Record Test")  # Set window title
        self.geometry("400x300")  # Set window size
        self.configure(bg=BG_COLOR)  # Set background color

        self.test_type = test_type  # Store the type of test (OneMinute or Stopwatch)

        self.setup_ui()  # Setup the user interface

    # Method to setup the UI of the record test window
    def setup_ui(self):
        tk.Label(self, text="Ready, Set, ...Type!", bg=BG_COLOR, fg=TEXT_COLOR, font=("Helvetica", 16, "bold")).pack(
            pady=10)

        tk.Label(self, text="Type The Data As It Was Given After You Took The Test", bg=BG_COLOR).pack(pady=5)

        # Labels and Entry boxes
        self.date_label = tk.Label(self, text="Date (MM/DD/YYYY):", bg=BG_COLOR)
        self.date_label.pack()
        self.date_entry = tk.Entry(self)
        self.date_entry.pack()

        if self.test_type == "Stopwatch":
            self.time_label = tk.Label(self, text="Stopwatch Time (HH:MM:SS):", bg=BG_COLOR)
            self.time_label.pack()
            self.time_entry = tk.Entry(self)
            self.time_entry.pack()

        self.words_label = tk.Label(self, text="Number of Words Typed:", bg=BG_COLOR)
        self.words_label.pack()
        self.words_entry = tk.Entry(self)
        self.words_entry.pack()

        self.chars_label = tk.Label(self, text="Number of Characters Typed:", bg=BG_COLOR)
        self.chars_label.pack()
        self.chars_entry = tk.Entry(self)
        self.chars_entry.pack()

        self.chars_per_sec_label = tk.Label(self, text="Number of Characters per Second:", bg=BG_COLOR)
        self.chars_per_sec_label.pack()
        self.chars_per_sec_entry = tk.Entry(self)
        self.chars_per_sec_entry.pack()

        # Button to add data
        add_button = tk.Button(self, text="Add", command=self.add_to_excel, bg=BUTTON_COLOR, fg=TEXT_COLOR)
        add_button.pack(pady=10)

        # Button to view results
        view_button = tk.Button(self, text="View Results", command=self.view_results, bg=BUTTON_COLOR, fg=TEXT_COLOR)
        view_button.pack()

    # Method to add test results to Excel
    def add_to_excel(self):
        try:
            # Get values from entry boxes
            date = self.date_entry.get().strip()
            words = int(self.words_entry.get().strip())
            chars = int(self.chars_entry.get().strip())
            chars_per_sec = float(self.chars_per_sec_entry.get().strip())

            if self.test_type == "Stopwatch":
                time = self.time_entry.get().strip()

                # Validate time format
                if len(time.split(':')) != 3:
                    raise ValueError("Invalid time format. Use HH:MM:SS.")

            # Validate date format
            if len(date.split('/')) != 3:
                raise ValueError("Invalid date format. Use MM/DD/YYYY.")

            # Validate integers
            if not (isinstance(words, int) and isinstance(chars, int)):
                raise ValueError("Words and Characters must be integers.")

            # Validate float
            if not isinstance(chars_per_sec, float):
                raise ValueError("Characters per Second must be a float.")

            # Create or load workbook and add data
            if self.test_type == "OneMinute":
                filename = "One_Minute_Test_Results.xlsx"
                headers = ['Date', 'Number of Words Typed', 'Number of Characters Typed',
                           'Number of Characters per Second']
            else:
                filename = "Stopwatch_Timing_Test.xlsx"
                headers = ['Date', 'Stopwatch Time', 'Number of Words Typed', 'Number of Characters Typed',
                           'Number of Characters per Second']

            wb = openpyxl.Workbook()

            # Check if sheet exists, otherwise create it
            if "Sheet" in wb.sheetnames:
                ws = wb["Sheet"]
            else:
                ws = wb.active

            # Write headers if the sheet is new
            if ws.max_row == 1:
                ws.append(headers)

            # Prepare data and append to the sheet
            if self.test_type == "Stopwatch":
                data = [date, time, words, chars, chars_per_sec]
            else:
                data = [date, words, chars, chars_per_sec]

            ws.append(data)

            # Save the workbook
            wb.save(filename)
            messagebox.showinfo("Success", "Data added successfully!")
            self.destroy()  # Close the record window after successful data addition

        except ValueError as e:
            messagebox.showerror("Error", str(e))
            return

    # Method to view results from the Excel file
    def view_results(self):
        if self.test_type == "OneMinute":
            filename = "One_Minute_Test_Results.xlsx"
        elif self.test_type == "Stopwatch":
            filename = "Stopwatch_Timing_Test.xlsx"

        try:
            os.startfile(filename)  # Open the Excel file
        except FileNotFoundError:
            messagebox.showerror("Error", f"No results found in {filename}")


# Entry point of the program
if __name__ == "__main__":
    app = TypingTestApp()  # Create an instance of the main application
    app.mainloop()  # Start the tkinter main loop

"""Explanation:
Imports: Import necessary libraries and modules for GUI (tkinter), file operations (os), Excel handling (openpyxl), and components from custom modules (typing_test, stopwatch_test).

Constants: Define color constants used for GUI elements.

TypingTestApp Class: Defines the main application window (TypingTestApp) with buttons to start various tests (One-Minute Typing Test and Stopwatch Timing Test) and record their results.

RecordTestWindow Class: This class represents a window (Toplevel) for recording test results. It allows users to enter data such as date, time (for Stopwatch test), words typed, characters typed, and characters per second. It validates the input data format and stores the results in Excel files (One_Minute_Test_Results.xlsx or `Stop"""




