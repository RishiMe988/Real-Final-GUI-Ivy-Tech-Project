import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import time  # Required for timer functionality

# Constants for colors and styling
BG_COLOR = "#D3BDB0"
BUTTON_COLOR = "#89937C"
TEXT_COLOR = "white"

# Class for the typing test window
class TypingTestWindow(tk.Toplevel):
    """Window for conducting a one-minute typing test."""

    def __init__(self, master):
        """Initialize the TypingTestWindow."""
        super().__init__(master)
        self.title("One-Minute Typing Test")
        self.geometry("400x300")
        self.configure(bg=BG_COLOR)

        # Timer label
        self.timer_label = tk.Label(self, text="Timer: 1:00", bg=BG_COLOR, fg=TEXT_COLOR)
        self.timer_label.pack(pady=10)

        # Text box for typing
        self.text_box = tk.Text(self, height=10, width=50)
        self.text_box.pack(pady=10)

        # Start button to begin the typing test
        self.start_button = tk.Button(self, text="Start", command=self.start_typing_test, bg=BUTTON_COLOR, fg=TEXT_COLOR)
        self.start_button.pack()

        # Go back button to close the window
        self.go_back_button = tk.Button(self, text="Go Back", command=self.destroy, bg=BUTTON_COLOR, fg=TEXT_COLOR)
        self.go_back_button.pack(side=tk.BOTTOM, padx=10, pady=10)

    def start_typing_test(self):
        """Starts the typing test and disables the start button."""
        self.start_button.config(state=tk.DISABLED)
        self.start_time = time.time()  # Record the start time
        self.update_timer()  # Begin updating the timer

    def update_timer(self):
        """Updates the timer display every 100ms."""
        elapsed_time = time.time() - self.start_time
        remaining_time = max(0, 60 - elapsed_time)
        minutes = int(remaining_time // 60)
        seconds = int(remaining_time % 60)
        self.timer_label.config(text=f"Timer: {minutes}:{seconds:02}")

        if remaining_time > 0:
            self.after(100, self.update_timer)  # Schedule the next update after 100ms
        else:
            self.calculate_results()  # If time is up, calculate and display results

    def calculate_results(self):
        """Calculates typing test results and displays them in a message box."""
        typed_text = self.text_box.get("1.0", tk.END)
        words = len(typed_text.split())
        characters = len(typed_text.replace(" ", "").replace("\n", ""))
        time_elapsed = time.time() - self.start_time
        characters_per_sec = characters / time_elapsed if time_elapsed > 0 else 0

        result_message = f"Words typed: {words}\nCharacters typed: {characters}\nCharacters per second: {characters_per_sec:.2f}"
        messagebox.showinfo("Typing Test Results", result_message)
        self.start_button.config(state=tk.NORMAL)  # Enable start button again
        self.text_box.delete("1.0", tk.END)  # Clear the text box for next test

# Class for recording typing test results to Excel
class RecordTypingTestWindow(tk.Toplevel):
    """Window for recording typing test results to Excel."""

    def __init__(self, master):
        """Initialize the RecordTypingTestWindow."""
        super().__init__(master)
        self.title("Record Typing Test Results")
        self.geometry("400x300")
        self.configure(bg=BG_COLOR)

        # Labels and entry fields for date, words, characters, and chars per sec
        self.date_label = tk.Label(self, text="Date (MM/DD/YYYY):", bg=BG_COLOR, fg=TEXT_COLOR)
        self.date_label.pack()
        self.date_entry = tk.Entry(self)
        self.date_entry.pack()

        self.words_label = tk.Label(self, text="Number of Words Typed:", bg=BG_COLOR, fg=TEXT_COLOR)
        self.words_label.pack()
        self.words_entry = tk.Entry(self)
        self.words_entry.pack()

        self.characters_label = tk.Label(self, text="Number of Characters Typed:", bg=BG_COLOR, fg=TEXT_COLOR)
        self.characters_label.pack()
        self.characters_entry = tk.Entry(self)
        self.characters_entry.pack()

        self.chars_per_sec_label = tk.Label(self, text="Characters per Second:", bg=BG_COLOR, fg=TEXT_COLOR)
        self.chars_per_sec_label.pack()
        self.chars_per_sec_entry = tk.Entry(self)
        self.chars_per_sec_entry.pack()

        # Buttons to add data to Excel and view results in Excel
        self.add_button = tk.Button(self, text="Add", command=self.add_to_excel, bg=BUTTON_COLOR, fg=TEXT_COLOR)
        self.add_button.pack()

        self.view_results_button = tk.Button(self, text="View Results", command=self.view_results, bg=BUTTON_COLOR, fg=TEXT_COLOR)
        self.view_results_button.pack()

    def add_to_excel(self):
        """Validates input and adds typing test results to Excel."""
        try:
            datetime.strptime(self.date_entry.get(), '%m/%d/%Y')
            int(self.words_entry.get())
            int(self.characters_entry.get())
            float(self.chars_per_sec_entry.get())

            # Load existing workbook or create a new one
            wb = openpyxl.load_workbook("One_Minute_Test_Results.xlsx")
            ws = wb.active
            if "Typing Test Results" not in wb.sheetnames:
                ws = wb.create_sheet("Typing Test Results")
                headers = ["Date", "Words Typed", "Characters Typed", "Characters per Second"]
                ws.append(headers)
                for col_num, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = cell.font.copy(bold=True)

            # Add data to the worksheet
            data = [
                self.date_entry.get(),
                int(self.words_entry.get()),
                int(self.characters_entry.get()),
                float(self.chars_per_sec_entry.get())
            ]
            ws.append(data)
            wb.save("One_Minute_Test_Results.xlsx")
            messagebox.showinfo("Success", "Data added to Excel!")

        except ValueError:
            messagebox.showerror("Error", "Invalid input data format!")

    def view_results(self):
        """Opens Excel to view the typing test results."""
        import os
        os.system("start excel.exe One_Minute_Test_Results.xlsx")

# Main application class
class TypingTestApp(tk.Tk):
    """Main application window for the Typing Test App."""

    def __init__(self):
        """Initialize the TypingTestApp."""
        super().__init__()
        self.title("Typing Test App")
        self.geometry("400x200")
        self.configure(bg=BG_COLOR)

        # Buttons to open typing test and record typing test windows
        button1 = tk.Button(self, text="One-Minute Typing Test", command=self.open_typing_test, bg=BUTTON_COLOR, fg=TEXT_COLOR)
        button1.pack(pady=20)

        button2 = tk.Button(self, text="Record One-Minute Typing Test", command=self.open_record_typing_test, bg=BUTTON_COLOR, fg=TEXT_COLOR)
        button2.pack()

    def open_typing_test(self):
        """Opens the one-minute typing test window."""
        typing_test_app = TypingTestWindow(self)
        typing_test_app.mainloop()

    def open_record_typing_test(self):
        """Opens the record typing test results window."""
        record_typing_test_app = RecordTypingTestWindow(self)
        record_typing_test_app.mainloop()

# Main entry point of the application
if __name__ == "__main__":
    app = TypingTestApp()
    app.mainloop()